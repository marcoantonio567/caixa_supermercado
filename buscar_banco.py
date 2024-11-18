import openpyxl
import re

def get_all_products(aba_produtos):
    planilha_produtos = openpyxl.load_workbook('produtos.xlsx')
    pagina_produtos = planilha_produtos[aba_produtos]

    for linha in pagina_produtos.iter_rows(values_only=True, min_row=2):
        codigo_produto, nome_produto, valor_produto, validade, quantidade, status = linha
        print(f"Código: {codigo_produto}")
        print(f"Nome: {nome_produto}")
        print(f"Valor: {valor_produto}")
        print(f"Validade: {validade}")
        print(f"Quantidade: {quantidade}")
        print(f"Status: {status}")
        print('-' * 40)  # Para separar visualmente cada produto
def mostrar_todos_produtos():
    get_all_products('Produtos Normais')
    get_all_products('Produtos Eletrônicos')
def extrair_dias(validade):
    # Usamos regex para capturar o número antes de "dias"
    match = re.match(r"(\d+)\s*dia[s]?", validade)
    if match:
        return int(match.group(1))
    return 0  # Retorna 0 caso não consiga extrair o número
def filtrar_produtos(aba_produtos, nome=None, quantidade=None, validade=None, codigo=None, status_ativo=True):
    planilha_produtos = openpyxl.load_workbook('produtos.xlsx')
    pagina_produtos = planilha_produtos[aba_produtos]

    encontrou = False

    for linha in pagina_produtos.iter_rows(values_only=True, min_row=2):
        codigo_produto, nome_produto, valor_produto, validade_produto, quantidade_produto, status = linha

        # Aplicar filtros apenas se os critérios forem fornecidos
        if codigo is not None and codigo.lower() != codigo_produto.lower():
            continue
        if nome is not None and nome.lower() not in nome_produto.lower():
            continue
        if quantidade is not None and quantidade <= quantidade_produto:
            continue
        if validade is not None:
            dias_validade = extrair_dias(validade_produto)
            if dias_validade > validade:
                continue
        if status_ativo and status.lower() != 'ativo':
            continue

        # Exibe os produtos que atendem aos critérios
        print(f"Código: {codigo_produto}")
        print(f"Nome: {nome_produto}")
        print(f"Valor: {valor_produto}")
        print(f"Validade: {validade_produto}")
        print(f"Quantidade: {quantidade_produto}")
        print(f"Status: {status}")
        print('-' * 40)  # Para separar visualmente cada produto
        encontrou = True
    if not encontrou:
        print("nenhum produto encontrado nessa aba")
        print("")
def mostrar_produtos_filtrados(codigo=None, nome=None, quantidade=None, validade=None):
    print("Produtos Normais:")
    print("")
    filtrar_produtos('Produtos Normais', nome, quantidade, validade, codigo)
    print("Produtos Eletrônicos:")
    print("")
    filtrar_produtos('Produtos Eletrônicos', nome, quantidade, validade, codigo)
def adicionar_produto(aba_produtos, codigo, nome, valor, validade, quantidade, status):
    planilha_produtos = openpyxl.load_workbook('produtos.xlsx')
    pagina_produtos = planilha_produtos[aba_produtos]

    # Adiciona o novo produto na primeira linha disponível
    nova_linha = [codigo, nome, valor, validade, quantidade, status]
    pagina_produtos.append(nova_linha)

    # Salva a planilha com as alterações
    planilha_produtos.save('produtos.xlsx')
    print(f"Produto '{nome}' adicionado na aba '{aba_produtos}'.")
def desativar_produto(codigo_produto):
    planilha_produtos = openpyxl.load_workbook('produtos.xlsx')
    
    # Função interna para atualizar o status do produto na aba fornecida
    def atualizar_status_aba(aba_produtos):
        pagina_produtos = planilha_produtos[aba_produtos]
        for linha in pagina_produtos.iter_rows(values_only=False, min_row=2):  # Iterar pelas células
            codigo_cell, nome_cell, valor_cell, validade_cell, quantidade_cell, status_cell = linha

            if codigo_cell.value == codigo_produto and status_cell.value.lower() == 'ativo':
                # Confirmar se o usuário deseja desativar
                resposta = input(f"O produto '{nome_cell.value}' está ativo. Deseja desativá-lo? (s/n): ").strip().lower()
                if resposta == 's':
                    # Altera o status do produto para 'Desativo'
                    status_cell.value = 'Desativo'
                    print(f"Produto '{nome_cell.value}' desativado com sucesso!")
                    return True
                else:
                    print("Operação cancelada.")
                    return False
        return False  # Caso o produto não seja encontrado ou já esteja desativado

    # Atualiza nas duas abas: Produtos Normais e Produtos Eletrônicos
    if atualizar_status_aba('Produtos Normais') or atualizar_status_aba('Produtos Eletrônicos'):
        # Salva a planilha com as alterações
        planilha_produtos.save('produtos.xlsx')
    else:
        print("Produto não encontrado ou já desativado.")
